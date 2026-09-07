"""Importa docentes da Escola Bahiana de Medicina (EBMSP)."""

import argparse
import asyncio
import csv
import json
import re
import sys
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import UUID

import httpx
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine
from unidecode import unidecode

ROOT = Path(__file__).resolve().parents[2]
INSTITUTION_ACRONYM = 'EBMSP'
DEFAULT_TERRITORIES = ROOT / 'storage/powerBI/dim_territorio_identidade.csv'
REPORT_PATH = ROOT / 'storage/teacher_import_report.csv'
MAX_WEEKLY_HOURS = 168
CEP_LENGTH = 8
ALIASES = {
    'nome_docente': 'name',
    'nome': 'name',
    'ch_semanal': 'carga_horaria',
    'carga_horaria_semanal': 'carga_horaria',
    'workload_hours_weekly': 'carga_horaria',
    'cep': 'zip_code',
    'territorio_de_identidade': 'territorio_identidade',
}


def normalized(value) -> str:
    return ' '.join(unidecode(str(value or '')).casefold().split())


def cell_text(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_rows(path: Path) -> list[dict]:
    if path.suffix.lower() == '.csv':
        for encoding in ('utf-8-sig', 'cp1252'):
            try:
                with path.open(encoding=encoding, newline='') as stream:
                    sample = stream.read(8192)
                    if not sample.strip():
                        return []
                    dialect = csv.Sniffer().sniff(sample, delimiters=';,\t')
                    stream.seek(0)
                    return list(csv.DictReader(stream, dialect=dialect))
            except UnicodeDecodeError:
                continue
        raise ValueError('Nao foi possivel ler o encoding do CSV.')
    if path.suffix.lower() != '.xlsx':
        raise ValueError('Use um arquivo .csv ou Excel .xlsx.')
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        values = worksheet.iter_rows(values_only=True)
        headers = next(values, ())
        return [dict(zip(headers, row, strict=True)) for row in values]
    finally:
        workbook.close()


def normalize_row(row: dict) -> dict:
    result = {}
    for key, value in row.items():
        if key is None:
            raise ValueError('Linha com mais valores que colunas.')
        header = re.sub(r'[^a-z0-9]+', '_', normalized(key)).strip('_')
        header = ALIASES.get(header, header)
        if header in result:
            raise ValueError(f'Coluna repetida: {header}')
        result[header] = cell_text(value)
    return result


def parse_workload(value: str) -> Decimal | None:
    if not value:
        return None
    try:
        hours = Decimal(value.replace(',', '.'))
    except InvalidOperation as exc:
        raise ValueError('Carga horaria invalida.') from exc
    if (
        not hours.is_finite()
        or not 0 <= hours <= MAX_WEEKLY_HOURS
        or hours != hours.quantize(Decimal('0.01'))
    ):
        raise ValueError(
            'Carga horaria deve estar entre 0 e 168 (2 decimais).'
        )
    return hours


def normalize_cep(value: str) -> str | None:
    if not value:
        return None
    digits = re.sub(r'[.\s-]', '', value)
    # Excel/CSV exports may omit a leading zero in numeric postal codes.
    if digits.isascii() and digits.isdigit():
        digits = digits.zfill(CEP_LENGTH)
    if not re.fullmatch(r'[0-9]{8}', digits):
        raise ValueError('CEP invalido: esperado um codigo de 8 digitos.')
    # A short number is not a recoverable leading-zero export.
    if len(re.sub(r'[.\s-]', '', value)) < CEP_LENGTH - 1:
        raise ValueError('CEP incompleto.')
    return digits


def load_territories(path: Path) -> dict[str, str]:
    result = {}
    for raw in read_rows(path):
        row = normalize_row(raw)
        city = normalized(row.get('municipio'))
        territory = row.get('territorio')
        if not city or not territory:
            raise ValueError('Mapa deve conter Municipio e Territorio.')
        if city in result and result[city] != territory:
            raise ValueError('Municipio com territorios conflitantes no mapa.')
        result[city] = territory
    return result


class CepResolver:
    def __init__(self, client: httpx.AsyncClient):
        self.client = client
        self.cache = {}

    async def fetch(self, cep: str) -> dict:
        response = await self.client.get(
            f'https://viacep.com.br/ws/{cep}/json/'
        )
        response.raise_for_status()
        data = response.json()
        if not isinstance(data, dict) or data.get('erro'):
            raise ValueError('CEP nao encontrado ou resposta invalida.')
        if (
            not isinstance(data.get('localidade'), str)
            or not data['localidade'].strip()
            or not isinstance(data.get('uf'), str)
            or not re.fullmatch(r'[A-Z]{2}', data['uf'])
        ):
            raise ValueError('CEP nao encontrado ou resposta invalida.')
        return data

    async def resolve(self, cep: str) -> dict:
        if cep not in self.cache:
            # Only the postal code is sent to ViaCEP, never the input row.
            try:
                self.cache[cep] = await self.fetch(cep)
            except (httpx.HTTPError, ValueError) as exc:
                self.cache[cep] = ValueError(
                    'Nao foi possivel resolver o CEP.'
                )
                raise self.cache[cep] from exc
            finally:
                await asyncio.sleep(0.2)
        if isinstance(self.cache[cep], Exception):
            raise self.cache[cep]
        return self.cache[cep]


async def resolve_institution(session):
    rows = (
        (
            await session.execute(
                text('SELECT id, name, acronym FROM institution')
            )
        )
        .mappings()
        .all()
    )
    matches = [
        row
        for row in rows
        if normalized(row['acronym']) == normalized(INSTITUTION_ACRONYM)
    ]
    if len(matches) != 1:
        raise ValueError(
            'Cadastre a Escola Bahiana de Medicina com a sigla EBMSP '
            'antes de executar este importador.'
        )
    return matches[0]['id']


# Integracao futura CPF -> Lattes. Manter desativada ate configurar o servico.
# CNPQ_CPF_ENDPOINT = ''  # Preencher endpoint e autenticacao autorizados.
#
# async def buscar_lattes_por_cpf(cpf: str) -> str:
#     cpf = re.sub(r'[^0-9]', '', cpf).zfill(11)
#     if len(cpf) != 11:
#         raise ValueError('CPF invalido.')
#     async with httpx.AsyncClient(timeout=30) as client:
#         response = await client.get(CNPQ_CPF_ENDPOINT, params={'cpf': cpf})
#         response.raise_for_status()
#         # Ajustar a extracao ao contrato real da resposta do CNPq.
#         lattes_id = str(response.json()['lattes_id']).strip()
#     if not re.fullmatch(r'[0-9]{16}', lattes_id):
#         raise ValueError('CNPq nao retornou um identificador Lattes valido.')
#     return lattes_id


def researcher_indexes(researchers: list) -> dict:
    indexes = {
        key: defaultdict(list)
        for key in ('researcher_id', 'lattes_id', 'name')
    }
    for researcher in researchers:
        indexes['researcher_id'][str(researcher['id'])].append(researcher)
        indexes['lattes_id'][researcher['lattes_id']].append(researcher)
        indexes['name'][normalized(researcher['name'])].append(researcher)
    return indexes


def match_researcher(row: dict, indexes: dict) -> dict:
    for key in ('researcher_id', 'lattes_id', 'name'):
        value = row.get(key)
        if not value:
            continue
        if key == 'researcher_id':
            value = str(UUID(value))
        if key == 'name':
            value = normalized(value)
        matches = indexes[key].get(value, [])
        if len(matches) != 1:
            raise ValueError(
                'Pesquisador nao encontrado ou identificacao ambigua.'
            )
        candidate = matches[0]
        if row.get('lattes_id') and candidate['lattes_id'] != row['lattes_id']:
            raise ValueError('Identificadores do pesquisador conflitantes.')
        if row.get('name') and normalized(candidate['name']) != normalized(
            row['name']
        ):
            raise ValueError(
                'Nome nao corresponde ao identificador informado.'
            )
        return candidate
    raise ValueError('Informe NOME_DOCENTE, lattes_id ou researcher_id.')


async def resolve_city(session, address: dict):
    rows = (
        (
            await session.execute(
                text("""
        SELECT c.id, c.name, c.country_id, s.abbreviation
        FROM city c LEFT JOIN state s ON s.id = c.state_id
        JOIN country country ON country.id = c.country_id
        WHERE upper(s.abbreviation) = :uf
            OR (c.state_id IS NULL AND upper(country.alpha_2_code) = 'BR')
    """),
                {'uf': address['uf']},
            )
        )
        .mappings()
        .all()
    )
    matches = [
        row
        for row in rows
        if normalized(row['name']) == normalized(address['localidade'])
    ]
    if len(matches) != 1:
        raise ValueError('Cidade/UF nao encontrada ou ambigua no banco.')
    return matches[0]


async def apply_record(session, record: dict, institution_id):
    await session.execute(
        text("""
        INSERT INTO researcher_institution (
            researcher_id, institution_id, territorio_identidade, carga_horaria
        ) VALUES (:rid, :iid, :territory, :hours)
        ON CONFLICT (researcher_id, institution_id) DO UPDATE SET
            territorio_identidade = CASE WHEN :update_territory
                THEN EXCLUDED.territorio_identidade
                ELSE researcher_institution.territorio_identidade END,
            carga_horaria = COALESCE(
                EXCLUDED.carga_horaria, researcher_institution.carga_horaria
            )
    """),
        {
            'rid': record['researcher_id'],
            'iid': institution_id,
            'territory': record['territory'],
            'hours': record['hours'],
            'update_territory': record['update_territory'],
        },
    )
    await session.execute(
        text("""
        UPDATE researcher SET institution_id = :iid
        WHERE id = :rid AND institution_id IS NULL
    """),
        {'rid': record['researcher_id'], 'iid': institution_id},
    )
    if record['city']:
        await session.execute(
            text("""
            UPDATE researcher SET city_id = :city_id, country_id = :country_id
            WHERE id = :rid
        """),
            {
                'rid': record['researcher_id'],
                'city_id': record['city']['id'],
                'country_id': record['city']['country_id'],
            },
        )


def prepare_record(raw: dict, indexes: dict, institution_id) -> dict:
    row = normalize_row(raw)
    researcher = match_researcher(row, indexes)
    if researcher['institution_id'] not in {None, institution_id}:
        raise ValueError('Pesquisador vinculado a outra universidade.')
    return {
        'researcher_id': researcher['id'],
        'hours': parse_workload(row.get('carga_horaria', '')),
        'cep': normalize_cep(row.get('zip_code', '')),
        'territory': row.get('territorio_identidade') or None,
        'existing_territory': researcher.get('existing_territory'),
        'existing_hours': researcher.get('existing_hours'),
    }


async def enrich_record(session, record: dict, resolver, territories: dict):
    record['city'] = None
    record['update_territory'] = record['territory'] is not None
    if not record['cep']:
        return
    address = await resolver.resolve(record['cep'])
    record['city'] = await resolve_city(session, address)
    if not record['territory']:
        record['territory'] = (
            territories.get(normalized(address['localidade']))
            if address['uf'] == 'BA'
            else None
        )
        if address['uf'] == 'BA' and not record['territory']:
            raise ValueError('Municipio ausente do mapa de territorios.')
    record['update_territory'] = True


async def import_rows(  # noqa: PLR0913
    session,
    rows: list,
    institution_id,
    resolver: CepResolver,
    territories: dict,
    *,
    dry_run: bool = False,
) -> list[dict]:
    researchers = (
        (
            await session.execute(
                text("""
        SELECT r.id, r.name, r.lattes_id, r.institution_id,
            ri.territorio_identidade AS existing_territory,
            ri.carga_horaria AS existing_hours
        FROM researcher r
        LEFT JOIN researcher_institution ri ON ri.researcher_id = r.id
            AND ri.institution_id = :institution_id
    """),
                {'institution_id': institution_id},
            )
        )
        .mappings()
        .all()
    )
    indexes = researcher_indexes(researchers)
    report = []
    prepared = defaultdict(list)
    for number, raw in enumerate(rows, start=2):
        item = {
            'linha': number,
            'status': 'ignorado',
            'motivo': '',
            'researcher_id': '',
            'universidade': INSTITUTION_ACRONYM,
            'institution_id': str(institution_id),
            'territorio_identidade': None,
            'carga_horaria': None,
        }
        report.append(item)
        try:
            # Por enquanto, NOME_DOCENTE identifica o pesquisador.
            # Para trocar para CPF -> CNPq -> Lattes, ativar este bloco
            # junto com buscar_lattes_por_cpf acima. Falhas devem ignorar
            # a linha, sem voltar silenciosamente para a busca por nome.
            # raw = normalize_row(raw)
            # if not raw.get('nu_cpf'):
            #     raise ValueError('CPF ausente na planilha.')
            # raw['lattes_id'] = await buscar_lattes_por_cpf(raw['nu_cpf'])
            # raw.pop('name', None)
            # raw.pop('researcher_id', None)
            record = prepare_record(raw, indexes, institution_id)
            item['researcher_id'] = str(record['researcher_id'])
            prepared[record['researcher_id']].append((record, item))
        except ValueError as exc:
            item['motivo'] = str(exc)

    for entries in prepared.values():
        record, item = entries[0]
        if any(other != record for other, _ in entries[1:]):
            for _, entry in entries:
                entry['motivo'] = (
                    'Linhas conflitantes para o mesmo pesquisador.'
                )
            continue
        for _, duplicate in entries[1:]:
            duplicate['motivo'] = 'Linha duplicada.'
        try:
            await enrich_record(session, record, resolver, territories)
            if not dry_run:
                await apply_record(session, record, institution_id)
            item.update({
                'status': 'simulado' if dry_run else 'atualizado',
                'territorio_identidade': (
                    record['territory']
                    if record['update_territory']
                    else record['existing_territory']
                ),
                'carga_horaria': (
                    record['hours']
                    if record['hours'] is not None
                    else record['existing_hours']
                ),
            })
        except ValueError as exc:
            item['motivo'] = str(exc)
    # The caller owns the transaction; database failures roll back the batch.
    return report


def write_report(path: Path, report: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8-sig', newline='') as stream:
        writer = csv.DictWriter(
            stream,
            fieldnames=[
                'linha',
                'status',
                'motivo',
                'researcher_id',
                'universidade',
                'institution_id',
                'territorio_identidade',
                'carga_horaria',
            ],
            delimiter=';',
        )
        writer.writeheader()
        writer.writerows(report)


async def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--file', '-f', type=Path, required=True)
    parser.add_argument(
        '--dry-run', action='store_true', help='Consulta sem gravar no banco.'
    )
    args = parser.parse_args()
    if REPORT_PATH.resolve() == args.file.resolve():
        parser.error(
            'O relatorio nao pode sobrescrever um arquivo de entrada.'
        )
    rows = read_rows(args.file)
    territories = load_territories(DEFAULT_TERRITORIES)
    sys.path.insert(0, str(ROOT / 'src'))
    from simcc.core.settings import Settings  # noqa: PLC0415

    engine = create_async_engine(Settings().DATABASE_URL)
    try:
        async with (
            async_sessionmaker(engine).begin() as session,
            httpx.AsyncClient(timeout=15) as client,
        ):
            if not await session.scalar(
                text("SELECT to_regclass('public.researcher_institution')")
            ):
                raise ValueError(
                    'Tabela researcher_institution ausente. '
                    'Aplique a migration 42b8e719dc60 antes de importar.'
                )
            institution_id = await resolve_institution(session)
            report = await import_rows(
                session,
                rows,
                institution_id,
                CepResolver(client),
                territories,
                dry_run=args.dry_run,
            )
        write_report(REPORT_PATH, report)
        counts = {
            status: sum(r['status'] == status for r in report)
            for status in ('atualizado', 'simulado', 'ignorado')
        }
        print(json.dumps({'total': len(rows), **counts}, ensure_ascii=True))
        print(f'Relatorio: {REPORT_PATH}')
    finally:
        await engine.dispose()


if __name__ == '__main__':
    # Psycopg async requires SelectorEventLoop on Windows.
    loop_factory = (
        asyncio.SelectorEventLoop if sys.platform == 'win32' else None
    )
    asyncio.run(main(), loop_factory=loop_factory)
